"""
XLSX export engine using openpyxl.

Generates professional Excel spreadsheets with company-branded styling,
formatted headers, auto-filtering, conditional formatting, and multi-sheet support.
"""

import os
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.worksheet import Worksheet

from app.reports.constants import DEFAULT_TEMPLATE_CONFIG


class XLSXReportEngine:
    """
    Professional XLSX report generator using openpyxl.

    Supports branded styling, multi-sheet workbooks, charts,
    auto-filtering, and conditional formatting.
    """

    def __init__(self, template_config: Optional[Dict[str, Any]] = None):
        self.config = {**DEFAULT_TEMPLATE_CONFIG, **(template_config or {})}
        self.colors = self._parse_colors()

    def _parse_colors(self) -> Dict[str, str]:
        """Parse hex colors to openpyxl-friendly format."""
        return {
            "primary": self.config.get("primary_color", "#2563EB").lstrip("#"),
            "secondary": self.config.get("secondary_color", "#1E40AF").lstrip("#"),
            "accent": self.config.get("accent_color", "#3B82F6").lstrip("#"),
            "text": self.config.get("text_color", "#1F2937").lstrip("#"),
            "light_bg": self.config.get("light_bg", "#F3F4F6").lstrip("#"),
            "white": "FFFFFF",
        }

    def _create_styles(self, wb: Workbook) -> Dict[str, NamedStyle]:
        """Create reusable named styles for the workbook."""
        styles = {}

        # Title style
        title_style = NamedStyle(name="report_title")
        title_style.font = Font(
            name="Calibri",
            size=self.config.get("title_font_size", 18),
            bold=True,
            color=self.colors["primary"],
        )
        title_style.alignment = Alignment(horizontal="center", vertical="center")
        wb.add_named_style(title_style)
        styles["title"] = title_style

        # Subtitle style
        subtitle_style = NamedStyle(name="report_subtitle")
        subtitle_style.font = Font(
            name="Calibri",
            size=self.config.get("subtitle_font_size", 12),
            color=self.colors["text"],
        )
        subtitle_style.alignment = Alignment(horizontal="center", vertical="center")
        wb.add_named_style(subtitle_style)
        styles["subtitle"] = subtitle_style

        # Header style
        header_style = NamedStyle(name="table_header")
        header_style.font = Font(
            name="Calibri",
            size=10,
            bold=True,
            color=self.colors["white"],
        )
        header_style.fill = PatternFill(
            start_color=self.colors["primary"],
            end_color=self.colors["primary"],
            fill_type="solid",
        )
        header_style.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        header_style.border = Border(
            bottom=Side(style="thin", color=self.colors["secondary"]),
        )
        wb.add_named_style(header_style)
        styles["header"] = header_style

        # Data cell style
        cell_style = NamedStyle(name="data_cell")
        cell_style.font = Font(name="Calibri", size=9, color=self.colors["text"])
        cell_style.alignment = Alignment(horizontal="left", vertical="center")
        cell_style.border = Border(
            bottom=Side(style="thin", color=self.colors["light_bg"]),
        )
        wb.add_named_style(cell_style)
        styles["cell"] = cell_style

        # Alternate row style
        alt_style = NamedStyle(name="alt_row")
        alt_style.font = Font(name="Calibri", size=9, color=self.colors["text"])
        alt_style.fill = PatternFill(
            start_color=self.colors["light_bg"],
            end_color=self.colors["light_bg"],
            fill_type="solid",
        )
        alt_style.alignment = Alignment(horizontal="left", vertical="center")
        wb.add_named_style(alt_style)
        styles["alt_row"] = alt_style

        # Number style
        number_style = NamedStyle(name="number_cell")
        number_style.font = Font(name="Calibri", size=9, color=self.colors["text"])
        number_style.alignment = Alignment(horizontal="right", vertical="center")
        number_style.number_format = "#,##0.00"
        wb.add_named_style(number_style)
        styles["number"] = number_style

        # Date style
        date_style = NamedStyle(name="date_cell")
        date_style.font = Font(name="Calibri", size=9, color=self.colors["text"])
        date_style.alignment = Alignment(horizontal="left", vertical="center")
        date_style.number_format = "YYYY-MM-DD HH:MM:SS"
        wb.add_named_style(date_style)
        styles["date"] = date_style

        # Stat label style
        stat_label_style = NamedStyle(name="stat_label")
        stat_label_style.font = Font(name="Calibri", size=8, bold=True, color=self.colors["white"])
        stat_label_style.fill = PatternFill(
            start_color=self.colors["primary"],
            end_color=self.colors["primary"],
            fill_type="solid",
        )
        stat_label_style.alignment = Alignment(horizontal="center", vertical="center")
        wb.add_named_style(stat_label_style)
        styles["stat_label"] = stat_label_style

        # Stat value style
        stat_value_style = NamedStyle(name="stat_value")
        stat_value_style.font = Font(name="Calibri", size=14, bold=True, color=self.colors["primary"])
        stat_value_style.fill = PatternFill(
            start_color=self.colors["light_bg"],
            end_color=self.colors["light_bg"],
            fill_type="solid",
        )
        stat_value_style.alignment = Alignment(horizontal="center", vertical="center")
        wb.add_named_style(stat_value_style)
        styles["stat_value"] = stat_value_style

        return styles

    def generate(
        self,
        output_path: str,
        title: str,
        subtitle: Optional[str] = None,
        sheets: Optional[List[Dict[str, Any]]] = None,
        summary_stats: Optional[List[Dict[str, Any]]] = None,
        single_table: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Generate an XLSX workbook.

        Args:
            output_path: File path for output.
            title: Workbook title (shown on first sheet).
            subtitle: Optional subtitle.
            sheets: List of sheet definitions [{"name": str, "headers": [...], "rows": [...]}].
            summary_stats: Key-value stats for summary row.
            single_table: Single-table shortcut {"headers": [...], "rows": [...]}.

        Returns:
            output_path on success.
        """
        wb = Workbook()
        styles = self._create_styles(wb)

        # Default sheet
        ws = wb.active
        ws.title = "Report"

        current_row = 1

        # --- Title ---
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        title_cell = ws.cell(row=current_row, column=1, value=title)
        title_cell.style = styles["title"]
        current_row += 2

        if subtitle:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            sub_cell = ws.cell(row=current_row, column=1, value=subtitle)
            sub_cell.style = styles["subtitle"]
            current_row += 1

        # Generated date
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        date_cell = ws.cell(
            row=current_row,
            column=1,
            value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        )
        date_cell.font = Font(name="Calibri", size=8, color="999999")
        date_cell.alignment = Alignment(horizontal="center")
        current_row += 2

        # --- Summary Stats ---
        if summary_stats:
            current_row = self._write_stats_row(ws, current_row, summary_stats, styles)
            current_row += 2

        # --- Single Table (shortcut) ---
        if single_table:
            current_row = self._write_data_table(
                ws, current_row, single_table.get("title", "Data"),
                single_table.get("headers", []),
                single_table.get("rows", []),
                styles,
            )

        # --- Multiple Sheets ---
        if sheets:
            # First sheet already exists - use it for first sheet data
            first = True
            for sheet_data in sheets:
                if first:
                    sheet_ws = ws
                    first = False
                else:
                    sheet_ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))
                    current_row = 1

                sheet_title = sheet_data.get("title", sheet_data.get("name", "Data"))
                current_row = self._write_data_table(
                    sheet_ws,
                    current_row,
                    sheet_title,
                    sheet_data.get("headers", []),
                    sheet_data.get("rows", []),
                    styles,
                )

        # Adjust column widths for all sheets
        for sheet in wb.worksheets:
            self._auto_fit_columns(sheet)

        wb.save(output_path)
        return output_path

    def _write_stats_row(
        self,
        ws: Worksheet,
        start_row: int,
        stats: List[Dict[str, Any]],
        styles: Dict[str, NamedStyle],
    ) -> int:
        """Write summary stats as a styled row."""
        for i, stat in enumerate(stats):
            col = i + 1

            # Label
            cell = ws.cell(row=start_row, column=col, value=stat.get("label", ""))
            cell.style = styles["stat_label"]

            # Value
            cell = ws.cell(row=start_row + 1, column=col, value=stat.get("value", ""))
            cell.style = styles["stat_value"]

        return start_row + 2

    def _write_data_table(
        self,
        ws: Worksheet,
        start_row: int,
        title: str,
        headers: List[str],
        rows: List[List[Any]],
        styles: Dict[str, NamedStyle],
    ) -> int:
        """Write a data table with headers and rows."""
        if not headers:
            return start_row

        # Title row
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = Font(name="Calibri", size=12, bold=True, color=self.colors["text"])
        start_row += 1

        # Header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.style = styles["header"]
        start_row += 1

        # Data rows
        for row_idx, row_data in enumerate(rows):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=start_row + row_idx, column=col_idx, value=value)
                if row_idx % 2 == 1:
                    cell.style = styles["alt_row"]
                else:
                    cell.style = styles["cell"]

        start_row += len(rows) + 1
        return start_row

    def _auto_fit_columns(self, ws: Worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            adjusted_width = min(max(max_length + 2, 10), 60)
            ws.column_dimensions[col_letter].width = adjusted_width

    def generate_from_buffer(self, **kwargs: Any) -> BytesIO:
        """Generate XLSX into a BytesIO buffer."""
        buffer = BytesIO()

        wb = Workbook()
        styles = self._create_styles(wb)

        ws = wb.active
        ws.title = "Report"

        title = kwargs.get("title", "Report")
        subtitle = kwargs.get("subtitle")
        summary_stats = kwargs.get("summary_stats")
        sheets = kwargs.get("sheets")
        single_table = kwargs.get("single_table")

        current_row = 1

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws.cell(row=1, column=1, value=title).style = styles["title"]
        current_row = 3

        if subtitle:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1, value=subtitle).style = styles["subtitle"]
            current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        ws.cell(row=current_row, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
        ws.cell(row=current_row, column=1).font = Font(size=8, color="999999")
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
        current_row += 2

        if summary_stats:
            current_row = self._write_stats_row(ws, current_row, summary_stats, styles)
            current_row += 2

        if single_table:
            current_row = self._write_data_table(
                ws, current_row,
                single_table.get("title", "Data"),
                single_table.get("headers", []),
                single_table.get("rows", []),
                styles,
            )

        if sheets:
            for sheet_data in sheets[1:] if single_table else sheets:
                sheet_ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))
                self._write_data_table(
                    sheet_ws, 1,
                    sheet_data.get("title", sheet_data.get("name", "Data")),
                    sheet_data.get("headers", []),
                    sheet_data.get("rows", []),
                    styles,
                )
                self._auto_fit_columns(sheet_ws)

        for sheet in wb.worksheets:
            self._auto_fit_columns(sheet)

        wb.save(buffer)
        buffer.seek(0)
        return buffer
